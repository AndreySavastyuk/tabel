# -*- coding: utf-8 -*-
"""Готовит изолированную папку _demo: копирует реальные входные данные и
наполняет справочники синтетикой, чтобы прогнать ВЕСЬ конвейер v0.3 на
настоящих данных СКУД и увидеть, как заполняются аналитические листы."""
import os
import shutil

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DEMO = os.path.join(HERE, "_demo")
LEZ = os.path.join(DEMO, "ЛЭЗ")
os.makedirs(LEZ, exist_ok=True)

# 1) копируем реальные входы
for rel in ["StorK.csv", "SIGUR.xlsx"]:
    shutil.copy(os.path.join(HERE, rel), os.path.join(DEMO, rel))
shutil.copy(os.path.join(HERE, "ЛЭЗ", "lez.xlsx"), os.path.join(LEZ, "lez.xlsx"))
shutil.copy(os.path.join(HERE, "_backup", "Сотрудники.txt"), os.path.join(LEZ, "Сотрудники.txt"))

# 2) читаем список сотрудников
names = []
with open(os.path.join(LEZ, "Сотрудники.txt"), encoding="utf_8_sig") as fh:
    for line in fh:
        s = line.strip()
        if not s:
            continue
        s = s.replace("!-", "").replace("!", "")
        s = s.split("=")[0].strip()
        if s:
            names.append(s)
# уникальные, сохраняя порядок
seen = set()
uniq = [n for n in names if not (n in seen or seen.add(n))]

# 3) Справочник_сотрудников.xlsx: раскидываем по отделам, всем график 5x2
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["ФИО", "Отдел", "Кабинет", "График", "Фикс.время", "Контроль ЛЭЗ"])
depts = ["Цех 1", "Цех 2", "Офис"]
for i, n in enumerate(uniq):
    dept = depts[i % 3]
    # большие цеха дробим по кабинетам, офис — без кабинета
    cab = f"Каб {i % 5 + 1}" if dept.startswith("Цех") else None
    lez_ctrl = "да" if dept.startswith("Цех") else "нет"
    ws.append([n, dept, cab, "5x2", None, lez_ctrl])
wb.save(os.path.join(LEZ, "Справочник_сотрудников.xlsx"))

# 4) Графики_нормы.xlsx: норма 5x2 на апрель 2026 (данные за апрель)
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["График", "Месяц", "Норма", "Начало смены", "Длит.смены", "Обед нач", "Обед кон"])
ws.append(["5x2", "2026-04", 175, "08:00", 11, "12:00", "12:30"])
wb.save(os.path.join(LEZ, "Графики_нормы.xlsx"))

# 5) Отсутствия.xlsx + Командировки.xlsx на реальных ФИО
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["ФИО", "Тип", "Дата с", "Дата по"])
if len(uniq) > 5:
    ws.append([uniq[3], "отпуск", "01.04.2026", "30.04.2026"])   # весь месяц в отпуске
    ws.append([uniq[4], "больничный", "10.04.2026", "15.04.2026"])
wb.save(os.path.join(LEZ, "Отсутствия.xlsx"))

wb = openpyxl.Workbook()
ws = wb.active
ws.append(["ФИО", "Дата с", "Дата по"])
if len(uniq) > 7:
    ws.append([uniq[6], "05.04.2026", "09.04.2026"])
wb.save(os.path.join(LEZ, "Командировки.xlsx"))

print(f"_demo готов: {len(uniq)} сотрудников, отделы {depts}")
print("отпуск весь месяц:", uniq[3] if len(uniq) > 5 else "-")
