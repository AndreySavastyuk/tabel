# -*- coding: utf-8 -*-
"""Интеграционный тест Фазы 4 на синтетических справочниках:
свод периодов + аналитические листы (отделы/бухгалтерия/нормы/опоздания)."""
import os
import sys

import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import compute
import model
import refdata
import report


def day(date, worked=None, absence=None, late=0, ot=0.0, dn=8.0, lunch=0.0):
    dr = model.DayRecord(name="", date=date)
    if absence:
        dr.absence = absence
    else:
        dr.entry, dr.exit = "08:00", "17:00"
        dr.worked_hours = worked
        dr.lunch_deducted = lunch
    dr.day_norm = dn
    dr.lateness_min = late
    dr.overtime_h = ot
    return dr


ref = refdata.RefData()
for n, dept in [("Иванов", "Цех1"), ("Петров", "Офис"), ("Сидоров", "Цех1")]:
    ref.dept_by_name[n] = dept
    ref.schedule_by_name[n] = "5x2"
ref.norms[("5x2", "2026-04")] = 160
ref.shift_len["5x2"] = 8

records = {}
# Иванов: 8 раб.дней по 8ч + 5 дней отпуск -> 64 + 5*8=40 = 104 / 160 = 65%
ri = [day(f"{d:02d}.04.2026", worked=8.0) for d in range(1, 9)]
ri += [day(f"{d:02d}.04.2026", absence="отпуск") for d in range(9, 14)]
# Петров: 4 дня по 8ч = 32 / 160 = 20% (<25%), один день опоздание, один переработка
rp = [day(f"{d:02d}.04.2026", worked=8.0) for d in range(1, 5)]
rp[0].lateness_min = 20
rp[1].overtime_h = 1.5
# Сидоров: 22 дня по 8ч = 176 / 160 = 110%
rs = [day(f"{d:02d}.04.2026", worked=8.0) for d in range(1, 23)]
for r in ri:
    r.name, r.dept, r.cabinet = "Иванов", "Цех1", "Каб1"
for r in rp:
    r.name, r.dept = "Петров", "Офис"          # без кабинета
for r in rs:
    r.name, r.dept, r.cabinet = "Сидоров", "Цех1", "Каб2"
records = {"Иванов": ri, "Петров": rp, "Сидоров": rs}

periods = compute.build_employee_periods(records, ref=ref, months=["2026-04"])
assert abs(periods["Иванов"].percent - 65.0) < 0.1, periods["Иванов"].percent
assert periods["Иванов"].bucket == "50-75%", periods["Иванов"].bucket
assert abs(periods["Петров"].percent - 20.0) < 0.1, periods["Петров"].percent
assert periods["Петров"].bucket == "<25%", periods["Петров"].bucket
assert abs(periods["Сидоров"].percent - 110.0) < 0.1, periods["Сидоров"].percent
assert periods["Сидоров"].bucket == ">75%", periods["Сидоров"].bucket
assert periods["Петров"].late_count == 1 and periods["Петров"].late_minutes == 20
assert abs(periods["Петров"].overtime_total - 1.5) < 0.01
print("period aggregation: PASS")

out = os.path.join(HERE, "_test_phase4.xlsx")
empty = pd.DataFrame([[" "]], columns=["x"])
with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
    empty.to_excel(writer, sheet_name="_")
    wb = writer.book
    fmts = report.report_formats(wb)
    used = {"_"}
    dept_sheets = report.write_department_sheets(writer, records, fmts, used_names=used)
    report.write_accounting_sheet(writer, periods, fmts)
    report.write_norms_sheet(writer, periods, fmts)
    report.write_late_overtime_sheet(writer, periods, fmts)

wb2 = openpyxl.load_workbook(out)
names = wb2.sheetnames
# Цех1 разбит на два кабинета -> два отдельных листа
cab_sheets = [s for s in names if "Цех1" in s]
assert len(cab_sheets) == 2, cab_sheets
assert any("Офис" in s for s in names), names      # отдел без кабинета — один лист
assert "Бухгалтерия (аванс)" in names
assert "Нормы" in names
assert "Опоздания и переработки" in names
print("department/cabinet split:", [s for s in names if "Цех1" in s or "Офис" in s])

# бухгалтерия: первая строка данных — самый низкий % (Петров, <50%)
acc = wb2["Бухгалтерия (аванс)"]
first = [c.value for c in acc[2]]
assert first[0] == "Петров", first
# нормы: формула +/-
nrm = wb2["Нормы"]
has_formula = any(
    isinstance(c.value, str) and c.value.startswith("=F")
    for r in nrm.iter_rows(min_row=2) for c in r
)
assert has_formula, "ожидалась формула +/- в листе Нормы"
# опоздания: Петров присутствует
lo = wb2["Опоздания и переработки"]
lo_names = [r[0].value for r in lo.iter_rows(min_row=2)]
assert "Петров" in lo_names, lo_names
print("analytic sheets: PASS")
print("ALL PHASE 4 TESTS PASS")
