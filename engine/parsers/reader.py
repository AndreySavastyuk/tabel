# -*- coding: utf-8 -*-
"""Чтение сырых файлов без tkinter/логирования.

Воспроизводит формы данных FileEditor.Files.read_file:
  read_csv(path)  -> [ [cells...], ... ]            (как __csv_reader, ';')
  read_xlsx(path) -> { sheet: [ {A:.., B:..}, ... ] } | None   (как __xl_reader)
  read_old_xls(path) -> [ {A:.., B:..}, ... ]        (как old_xls_reader, Hikvision)
"""
import csv

import openpyxl
import pandas as pd

_LETTERS = [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P",
    "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z",
    "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM",
    "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
]


def read_csv(path):
    """Список строк CSV (разделитель ';'). None при ошибке.

    Кодировка задаётся ЯВНО (а не системной локалью, как в легаси open(path,'r')),
    чтобы движок не зависел от locale/PYTHONUTF8 на сервере. StorK.csv —
    cp1251; современные выгрузки бывают utf-8-sig. Поэтому пробуем utf-8-sig,
    затем откатываемся на cp1251 (что воспроизводит чтение легаси на Windows)."""
    for enc in ("utf_8_sig", "cp1251"):
        try:
            rows = []
            with open(path, "r", encoding=enc) as file:
                for row in csv.reader(file, delimiter=";"):
                    rows.append([i for i in row])
            return rows
        except UnicodeDecodeError:
            continue
        except Exception as e:
            print("read_csv:", e)
            return None
    print("read_csv: не удалось декодировать", path)
    return None


def read_xlsx(path):
    """{sheet: [dict_rows]} как FileEditor.__xl_reader, либо None при ошибке."""
    try:
        xl = pd.ExcelFile(path)
        wb_obj = openpyxl.load_workbook(path, data_only=True)
        data = {}
        for i in xl.sheet_names:
            rows = []
            sheet_obj = wb_obj[i]
            row = 1
            rowbroker = 100
            while rowbroker != 0:
                count = 1
                tmp = {}
                cellbroker = 10
                for cels in _LETTERS:
                    if sheet_obj.cell(row=row, column=count).value is None:
                        cellbroker -= 1
                    else:
                        cellbroker = 10
                    if cellbroker == 0:
                        for j in _LETTERS:
                            if tmp.get(j) is None and j in tmp.keys():
                                tmp.pop(j)
                        break
                    tmp[cels] = sheet_obj.cell(row=row, column=count).value
                    count += 1
                rows.append(tmp)
                if len(tmp) == 0:
                    rowbroker -= 1
                else:
                    rowbroker = 100
                if rowbroker == 0:
                    for _ in range(0, 100):
                        rows.pop(-1)
                row += 1
            data[i] = rows
        xl.close()
        return data
    except Exception as e:
        print("read_xlsx:", e)
        return None


def sheet_rows(xl_data):
    """Строки целевого листа из результата read_xlsx.

    В легаси FileEditor.__xl_reader из-за утечки переменной цикла лист
    индексировался ключом 99 (а парсеры читали data.get(99)) — фактически это
    строки ПОСЛЕДНЕГО листа книги. Здесь читаем корректно: последний лист.
    Возвращает список dict-строк или None."""
    if isinstance(xl_data, dict) and xl_data:
        return list(xl_data.values())[-1]
    return None


def read_old_xls(path):
    """Список dict-строк из HTML-подобного .xls (Hikvision). Перенос
    old_xls_reader без отладочных print."""
    with open(path, encoding="utf-8") as f:
        table = f.readlines()
        xml = ""
        for i in table:
            xml += str(i)
        xml = xml.replace("\n", "").replace("\'", "")
        data = str(xml).split('<table class="Detail2">')[-1].split("tr")
        trs = []
        for tr in data:
            if len(str(tr)) < 10:
                continue
            table = {}
            count = 0
            for td in tr.split("<td"):
                if len(str(td)) < 5 or "SECTION" in str(td) or "</body>" in td:
                    continue
                td = td.split(':@">')[1].replace("</td>", "")
                table[_LETTERS[count]] = td
                count += 1
                if count == 6:
                    break
            if table.get("F") is not None:
                trs.append(table)
        return trs
