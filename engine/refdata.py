# -*- coding: utf-8 -*-
"""Слой справочных данных (SCUD v0.3).

Читает справочники рядом с ЛЭЗ/Сотрудники.txt:
  * Справочник_сотрудников.xlsx — отдел, график, фикс.время, контроль ЛЭЗ
  * Графики_нормы.xlsx          — норма часов, смена и обед на (график × месяц)
  * Отсутствия.xlsx             — отпуск/больничный из бухгалтерии
  * Командировки.xlsx           — командировки (заполняет пользователь)

Всё устойчиво к отсутствию файлов: любой недостающий/битый источник просто
пропускается (per-source try/except), функция всегда возвращает RefData и
никогда не валит start().
"""
import json
import os
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

import openpyxl
import pandas as pd

from . import model

SETTINGS_FILE = "Настройки.json"

EMP_REF_FILE = "Справочник_сотрудников.xlsx"
EMP_REF_HEADERS = ["ФИО", "Отдел", "Кабинет", "График", "Фикс.время", "Контроль ЛЭЗ"]
NORMS_FILE = "Графики_нормы.xlsx"
NORMS_HEADERS = ["График", "Месяц", "Норма", "Начало смены", "Длит.смены", "Обед нач", "Обед кон"]
ABSENCE_FILE = "Отсутствия.xlsx"
ABSENCE_HEADERS = ["ФИО", "Тип", "Дата с", "Дата по"]
TRIP_FILE = "Командировки.xlsx"
TRIP_HEADERS = ["ФИО", "Дата с", "Дата по"]


@dataclass
class RefData:
    fixed_times: dict = field(default_factory=dict)       # name -> "HH:MM"
    dept_by_name: dict = field(default_factory=dict)      # name -> отдел
    cabinet_by_name: dict = field(default_factory=dict)   # name -> кабинет
    schedule_by_name: dict = field(default_factory=dict)  # name -> график
    lez_controlled: dict = field(default_factory=dict)    # name -> bool
    norms: dict = field(default_factory=dict)             # (график, "YYYY-MM") -> часы
    shift_start: dict = field(default_factory=dict)       # график -> "HH:MM"
    shift_len: dict = field(default_factory=dict)         # график -> часы (float)
    lunch: dict = field(default_factory=dict)             # график -> ("HH:MM", "HH:MM")
    absences: dict = field(default_factory=dict)          # name -> [(тип, date_from, date_to)]

    def dept(self, name):
        return self.dept_by_name.get(name) or "Без отдела"

    def cabinet(self, name):
        return self.cabinet_by_name.get(name) or None

    def schedule(self, name):
        return self.schedule_by_name.get(name)

    def is_lez_controlled(self, name):
        return bool(self.lez_controlled.get(name, False))

    def absence_on(self, name, d):
        """Тип отсутствия на дату d (datetime.date) или None."""
        for typ, d1, d2 in self.absences.get(name, []):
            if d1 <= d <= d2:
                return typ
        return None


# --------------------------------------------------------------------------
# Вспомогательные парсеры (терпимы к мусору / пустым ячейкам)
# --------------------------------------------------------------------------
def _norm_header(h):
    return str(h).strip().lower().replace("ё", "е").replace(".", "").replace(" ", "")


def _find_col(columns, *aliases):
    """Ищет колонку по подстроке среди нормализованных заголовков."""
    norm = {_norm_header(c): c for c in columns}
    for a in aliases:
        a = _norm_header(a)
        for nk, orig in norm.items():
            if a in nk:
                return orig
    return None


def _as_bool(v):
    s = str(v).strip().lower()
    return s in ("да", "yes", "1", "true", "y", "+", "да.")


def _as_time_str(v):
    """Любой ввод времени -> 'HH:MM' или None."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (datetime,)):
        return f"{v.hour:02d}:{v.minute:02d}"
    if hasattr(v, "hour") and hasattr(v, "minute"):  # datetime.time
        return f"{v.hour:02d}:{v.minute:02d}"
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return None
    if ":" in s:
        hh, mm = s.split(":")[:2]
        try:
            return f"{int(hh):02d}:{int(mm):02d}"
        except ValueError:
            return None
    return None


def _as_date(v):
    """Любой ввод даты -> datetime.date или None."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _as_float(v, default=0.0):
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return default
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return default


# --------------------------------------------------------------------------
# Загрузчики отдельных справочников
# --------------------------------------------------------------------------
def _load_employees_ref(path, ref, nf):
    df = pd.read_excel(path)
    cols = list(df.columns)
    c_fio = _find_col(cols, "фио", "сотрудник")
    c_dept = _find_col(cols, "отдел", "подразделение")
    c_cab = _find_col(cols, "кабинет", "каб")
    c_sched = _find_col(cols, "график")
    c_fix = _find_col(cols, "фикс")
    c_lez = _find_col(cols, "контрользэз", "лэз", "проходная")
    if c_fio is None:
        return
    for _, row in df.iterrows():
        name = nf(str(row[c_fio]).strip()) if not pd.isna(row[c_fio]) else None
        if not name:
            continue
        if c_dept and not pd.isna(row[c_dept]):
            ref.dept_by_name[name] = str(row[c_dept]).strip()
        if c_cab and not pd.isna(row[c_cab]):
            ref.cabinet_by_name[name] = str(row[c_cab]).strip()
        if c_sched and not pd.isna(row[c_sched]):
            ref.schedule_by_name[name] = str(row[c_sched]).strip()
        if c_fix:
            t = _as_time_str(row[c_fix])
            if t:
                ref.fixed_times[name] = t
        if c_lez:
            ref.lez_controlled[name] = _as_bool(row[c_lez])


def _load_norms_ref(path, ref):
    df = pd.read_excel(path)
    cols = list(df.columns)
    c_sched = _find_col(cols, "график")
    c_month = _find_col(cols, "месяц", "period", "период")
    c_norm = _find_col(cols, "норма")
    c_start = _find_col(cols, "началосмены", "начало")
    c_len = _find_col(cols, "длит", "продолжит")
    c_l1 = _find_col(cols, "обеднач", "обедс", "обедначало")
    c_l2 = _find_col(cols, "обедкон", "обедпо", "обедконец")
    if c_sched is None:
        return
    for _, row in df.iterrows():
        sched = str(row[c_sched]).strip() if not pd.isna(row[c_sched]) else None
        if not sched:
            continue
        if c_month and c_norm and not pd.isna(row[c_month]):
            mk = _month_key(row[c_month])
            if mk:
                ref.norms[(sched, mk)] = _as_float(row[c_norm])
        if c_start:
            t = _as_time_str(row[c_start])
            if t:
                ref.shift_start[sched] = t
        if c_len:
            v = _as_float(row[c_len], default=0.0)
            if v:
                ref.shift_len[sched] = v
        l1 = _as_time_str(row[c_l1]) if c_l1 else None
        l2 = _as_time_str(row[c_l2]) if c_l2 else None
        if l1 and l2:
            ref.lunch[sched] = (l1, l2)


def _month_key(v):
    """-> 'YYYY-MM' из даты/строки '2026-04' / '04.2026' / datetime."""
    if isinstance(v, (datetime, date)):
        return f"{v.year:04d}-{v.month:02d}"
    s = str(v).strip()
    if "-" in s:
        parts = s.split("-")
        if len(parts[0]) == 4:
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
    if "." in s:
        parts = s.split(".")
        if len(parts) == 2:
            return f"{int(parts[1]):04d}-{int(parts[0]):02d}"
    return None


def _load_absences(path, ref, nf, default_type=None):
    df = pd.read_excel(path)
    cols = list(df.columns)
    c_fio = _find_col(cols, "фио", "сотрудник")
    c_type = _find_col(cols, "тип", "вид", "причина")
    c_from = _find_col(cols, "датас", "сдата", "начало", "датаначала", "с")
    c_to = _find_col(cols, "датапо", "подата", "конец", "датаокончания", "по")
    if c_fio is None or c_from is None or c_to is None:
        return
    for _, row in df.iterrows():
        name = nf(str(row[c_fio]).strip()) if not pd.isna(row[c_fio]) else None
        d1 = _as_date(row[c_from])
        d2 = _as_date(row[c_to])
        if not name or not d1 or not d2:
            continue
        typ = default_type
        if c_type and not pd.isna(row[c_type]):
            typ = str(row[c_type]).strip()
        ref.absences.setdefault(name, []).append((typ or "отсутствие", d1, d2))


# --------------------------------------------------------------------------
# Точка входа
# --------------------------------------------------------------------------
def load_reference_data(wp, name_normalizer=None):
    """Собирает все справочники из <wp>/ЛЭЗ/. Никогда не падает."""
    nf = name_normalizer or (lambda x: x)
    base = os.path.join(wp, "ЛЭЗ")
    ref = RefData()

    sources = [
        ("Справочник_сотрудников.xlsx", lambda p: _load_employees_ref(p, ref, nf)),
        ("Графики_нормы.xlsx", lambda p: _load_norms_ref(p, ref)),
        ("Отсутствия.xlsx", lambda p: _load_absences(p, ref, nf)),
        ("Командировки.xlsx", lambda p: _load_absences(p, ref, nf, default_type="командировка")),
    ]
    for fname, loader in sources:
        path = os.path.join(base, fname)
        if not os.path.exists(path):
            continue
        try:
            loader(path)
        except Exception as e:
            print(f"load_reference_data: пропущен {fname}: {e}")

    print(
        "Справочники: отделы=%d, графики=%d, нормы=%d, отсутствия(чел)=%d"
        % (len(ref.dept_by_name), len(ref.schedule_by_name), len(ref.norms), len(ref.absences))
    )
    return ref


# --------------------------------------------------------------------------
# Авто-сопровождение справочника сотрудников + шаблоны
# --------------------------------------------------------------------------
def _header_index(headers, *aliases):
    for i, h in enumerate(headers):
        hn = _norm_header(h)
        for a in aliases:
            if _norm_header(a) in hn:
                return i
    return None


def sync_employee_reference(wp, names, fixed_times=None, name_normalizer=None):
    """Гарантирует, что в Справочник_сотрудников.xlsx есть строка под каждое
    ФИО. Существующие строки и заполненные пользователем колонки
    (Отдел/График/Контроль ЛЭЗ) НЕ затираются — только дописываются новые
    ФИО. Для новых строк мигрируется фикс. время из Сотрудники.txt."""
    nf = name_normalizer or (lambda x: x)
    fixed_times = fixed_times or {}
    base = os.path.join(wp, "ЛЭЗ")
    os.makedirs(base, exist_ok=True)
    path = os.path.join(base, EMP_REF_FILE)

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [(c.value if c.value is not None else "") for c in ws[1]]
        # Дополняем недостающие колонки-заголовки (например, новый «Кабинет»)
        # в конец первой строки — позиции колонок при чтении не важны.
        for h in EMP_REF_HEADERS:
            if _header_index(headers, h) is None:
                headers.append(h)
                ws.cell(row=1, column=len(headers), value=h)
        i_fio = _header_index(headers, "фио", "сотрудник")
        i_fix = _header_index(headers, "фикс")
        if i_fio is None:
            i_fio = 0
        width = max(len(headers), len(EMP_REF_HEADERS))
        seen = set()
        for row in ws.iter_rows(min_row=2):
            v = row[i_fio].value if i_fio < len(row) else None
            if v:
                seen.add(nf(str(v).strip()))
        added = 0
        for raw in names:
            key = nf(str(raw).strip())
            if not key or key in seen:
                continue
            seen.add(key)
            newrow = [None] * width
            newrow[i_fio] = raw
            if i_fix is not None and i_fix < width:
                newrow[i_fix] = fixed_times.get(key)
            ws.append(newrow)
            added += 1
        wb.save(path)
        print(f"Справочник сотрудников: +{added} новых ФИО (всего {len(seen)})")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Сотрудники"
        ws.append(EMP_REF_HEADERS)
        seen = set()
        for raw in names:
            key = nf(str(raw).strip())
            if not key or key in seen:
                continue
            seen.add(key)
            ws.append([raw, None, None, fixed_times.get(key), None])
        wb.save(path)
        print(f"Справочник сотрудников создан: {len(seen)} ФИО -> {path}")


def _create_xlsx(path, headers, rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in (rows or []):
        ws.append(r)
    wb.save(path)


def ensure_templates(wp):
    """Создаёт пустые шаблоны норм/отсутствий/командировок, если их ещё нет.
    Существующие файлы не трогает."""
    base = os.path.join(wp, "ЛЭЗ")
    os.makedirs(base, exist_ok=True)

    norms = os.path.join(base, NORMS_FILE)
    if not os.path.exists(norms):
        _create_xlsx(norms, NORMS_HEADERS,
                     [["ПРИМЕР_5x2", "2026-06", 151, "08:00", 8, "12:00", "12:30"]])
        print("Создан шаблон", norms)

    absc = os.path.join(base, ABSENCE_FILE)
    if not os.path.exists(absc):
        _create_xlsx(absc, ABSENCE_HEADERS)
        print("Создан шаблон", absc)

    trip = os.path.join(base, TRIP_FILE)
    if not os.path.exists(trip):
        _create_xlsx(trip, TRIP_HEADERS)
        print("Создан шаблон", trip)

    # Настройки: создаём, а если файл уже есть — дописываем НОВЫЕ пороги,
    # не затирая значения, которые пользователь уже поправил.
    settings = os.path.join(base, SETTINGS_FILE)
    data = {}
    if os.path.exists(settings):
        try:
            with open(settings, encoding="utf_8_sig") as fh:
                data = json.load(fh)
        except Exception:
            data = {}
    merged = dict(model.THRESHOLDS)
    merged.update(data)                       # значения пользователя в приоритете
    if merged != data:
        with open(settings, "w", encoding="utf_8_sig") as fh:
            json.dump(merged, fh, ensure_ascii=False, indent=2)
        print(("Создан" if not os.path.exists(settings) else "Обновлён"), "шаблон", settings)


def load_settings(wp):
    """Читает ЛЭЗ/Настройки.json и возвращает только известные пороги.
    Отсутствие файла / мусор -> {} (берутся значения по умолчанию)."""
    path = os.path.join(wp, "ЛЭЗ", SETTINGS_FILE)
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf_8_sig") as fh:
            data = json.load(fh)
        return {k: v for k, v in data.items() if k in model.THRESHOLDS}
    except Exception as e:
        print("load_settings:", e)
        return {}
