# -*- coding: utf-8 -*-
"""Проставляет колонку «Отдел» в Справочник_сотрудников.xlsx по списку
заказчика. Сопоставляет частичные имена (фамилия / фамилия+имя) с полными ФИО.

Запуск:  python назначить_отделы.py
Файл справочника при запуске должен быть ЗАКРЫТ в Excel.
Печатает отчёт: что не нашлось, что неоднозначно, кто в двух отделах.
"""
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)             # репозиторий (legacy/ лежит в корне)
PATH = os.path.join(ROOT, "ЛЭЗ", "Справочник_сотрудников.xlsx")

# Список «отдел -> [частичные ФИО]» — это персональные данные заказчика,
# поэтому он вынесен в отдельный файл `_departments_local.py`, который НЕ
# коммитится в репозиторий (см. .gitignore). Положите рядом свой
# `_departments_local.py` с переменной DEPARTMENTS того же формата.
# Если файла нет — используется обезличенный пример ниже.
try:
    from _departments_local import DEPARTMENTS
except ImportError:
    # (Отдел, [частичные имена]). Фамилия — точно, имя/отчество — с допуском.
    DEPARTMENTS = [
        ("Конструкторский отдел № 1", [
            "Иванов", "Петров Сергей", "Сидорова Анна"]),
        ("Цех сборки", [
            "Кузнецов", "Смирнов Денис", "Васильев Олег"]),
        ("Бухгалтерия", [
            "Морозова", "Новикова Анна"]),
    ]


def norm(s):
    return " ".join(str(s).lower().replace("ё", "е").split())


def tokens(s):
    return norm(s).split()


def lev(a, b):
    """Расстояние Левенштейна (для коротких строк)."""
    if a == b:
        return 0
    m, n = len(a), len(b)
    if abs(m - n) > 2:
        return 3
    prev = list(range(n + 1))
    for i in range(1, m + 1):
        cur = [i] + [0] * n
        for j in range(1, n + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            cur[j] = min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + cost)
        prev = cur
    return prev[n]


def sim(a, b):
    """Похожи ли токены (имя/отчество): общий 4-символьный префикс или опечатка."""
    if not a or not b:
        return False
    if a[:4] == b[:4]:
        return True
    return lev(a, b) <= 1


def fio_match(entry, fio_toks):
    """Совпадает ли запись списка с ФИО. Фамилия — точно (после ё→е); имя/
    отчество — с допуском на опечатки. Запись из 2 слов: 2-е слово может быть
    как именем, так и отчеством."""
    e, f = entry, fio_toks
    if not e or not f or e[0] != f[0]:
        return False
    if len(e) >= 3:
        if len(f) >= 3:
            return sim(e[1], f[1]) and sim(e[2], f[2])
        return len(f) >= 2 and sim(e[1], f[1])
    if len(e) == 2:
        return len(f) >= 2 and sim(e[1], f[1])     # фамилия + имя
    return True                      # только фамилия


def same_person(a_toks, b_toks):
    """Два ФИО — вероятно один человек (дубликат с опечаткой/ё/е)?"""
    if len(a_toks) >= 2 and len(b_toks) >= 2 and not sim(a_toks[1], b_toks[1]):
        return False
    if len(a_toks) >= 3 and len(b_toks) >= 3 and not sim(a_toks[2], b_toks[2]):
        return False
    return True


def main():
    wb = openpyxl.load_workbook(PATH)
    ws = wb.active
    headers = [(c.value if c.value is not None else "") for c in ws[1]]

    def col(*al):
        for i, h in enumerate(headers):
            hn = norm(h)
            for a in al:
                if norm(a) in hn:
                    return i
        return None

    i_fio = col("фио", "сотрудник")
    i_dept = col("отдел", "подразделение")
    if i_fio is None or i_dept is None:
        print("Не нашёл колонки ФИО/Отдел")
        return

    # ФИО из справочника
    fios = []
    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = row[i_fio].value if i_fio < len(row) else None
        if v:
            fios.append({"row": ri, "raw": str(v).strip(), "toks": tokens(v)})

    # порядок: сначала записи с именем (точнее), потом только фамилия
    entries = []
    for dept, members in DEPARTMENTS:
        for m in members:
            entries.append((dept, m, tokens(m)))
    entries.sort(key=lambda e: -len(e[2]))

    assignment = {}     # row -> (dept, entry)
    not_found, ambiguous, taken = [], [], []
    for dept, raw, toks in entries:
        cands_all = [f for f in fios if fio_match(toks, f["toks"])]
        cands = [f for f in cands_all if f["row"] not in assignment]
        if not cands:
            if cands_all:
                taken.append((dept, raw, sorted({assignment[f["row"]][0] for f in cands_all})))
            else:
                not_found.append((dept, raw))
        elif len(cands) == 1:
            assignment[cands[0]["row"]] = (dept, raw)
        elif all(same_person(cands[0]["toks"], c["toks"]) for c in cands[1:]):
            # несколько строк — это один человек (дубликаты): отделим оба
            for c in cands:
                assignment[c["row"]] = (dept, raw)
        else:
            ambiguous.append((dept, raw, [c["raw"] for c in cands]))

    # запись
    for f in fios:
        a = assignment.get(f["row"])
        ws.cell(row=f["row"], column=i_dept + 1, value=a[0] if a else None)

    out_path = PATH
    try:
        wb.save(PATH)
    except PermissionError:
        out_path = os.path.join(ROOT, "ЛЭЗ", "Справочник_сотрудников (с отделами).xlsx")
        wb.save(out_path)

    assigned_rows = set(assignment)
    unassigned = [f["raw"] for f in fios if f["row"] not in assigned_rows]

    rep = []
    rep.append(f"Записан файл: {out_path}")
    rep.append(f"Назначено отделов: {len(assignment)} из {len(fios)} ФИО")
    rep.append("")
    rep.append(f"[1] НЕ НАЙДЕНЫ в справочнике ({len(not_found)}) — нет такого ФИО среди отметок:")
    for dept, raw in not_found:
        rep.append(f"    «{raw}»  ({dept})")
    rep.append("")
    rep.append(f"[2] НЕОДНОЗНАЧНО ({len(ambiguous)}) — несколько однофамильцев, уточните имя:")
    for dept, raw, cands in ambiguous:
        rep.append(f"    «{raw}» ({dept}) -> {cands}")
    rep.append("")
    rep.append(f"[3] В ДВУХ ОТДЕЛАХ ({len(taken)}) — записан в первый, проверьте:")
    for dept, raw, where in taken:
        rep.append(f"    «{raw}»: список {dept}, но уже отдан в {where}")
    rep.append("")
    rep.append(f"[4] БЕЗ ОТДЕЛА ({len(unassigned)}) — есть в отметках, но нет в вашем списке:")
    for r in sorted(unassigned):
        rep.append(f"    {r}")

    text = "\n".join(rep)
    with open(os.path.join(ROOT, "отчет_отделы.txt"), "w", encoding="utf-8") as fh:
        fh.write(text)
    print(f"Назначено: {len(assignment)}/{len(fios)}. "
          f"Не найдено: {len(not_found)}, неоднозначно: {len(ambiguous)}, "
          f"в двух отделах: {len(taken)}, без отдела: {len(unassigned)}.")
    print("Подробности — в отчет_отделы.txt")


if __name__ == "__main__":
    main()
