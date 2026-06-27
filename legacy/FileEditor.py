import pandas as pd
import json
import openpyxl
import os
import csv
from time import sleep as sleep
from datetime import datetime, date, time
import tkinter.filedialog as fd
import tkinter


class Files:

    def __init__(self, **kwargs):
        self.logs = kwargs.get("logs", False)
        self.logtype=kwargs.get("logtype", "print")
        self._path_to_log = kwargs.get("path_to_logs", None)
        self.__path_to_file = None
        self._path_to_outputs = kwargs.get("path_to_outputs", None)
        self._file_encoding = "utf_8_sig"
        self.info = {}
        self.data = None


    def __str__(self):
        return {"path_to_file": self.__path_to_file,
                "path_to_log": self._path_to_log,
                "_path_to_outputs":  self._path_to_outputs,
                "file_encoding": self._file_encoding,
                "info": self.info}

    def __file_init(self):
        self.__loger(f'Файл взят в работу\n{self.info}')
        if str(self.info.get("file_type")).lower() == "csv":
            self.__loger(f'[CSV READER]')
            self.__csv_reader()
        if str(self.info.get("file_type")).lower() == "xls" or str(self.info.get("file_type")).lower() == "xlsx":
            self.__loger(f'[EXCEL READER]')
            self.__xl_reader()

    def file_selector(self):
        self.__loger("Пользовательский выбор файла")
        sleep(3)
        self.root = tkinter.Tk()
        self.__path_to_file = fd.askopenfilename(parent=self.root, title="Тыкни нужный фапйл", filetypes=[("Excel files (xls, xlsx, csv)", ".xlsx .xls .csv"), (f'Текстовые файлы (TXT)', "*.txt")])
        self.__dir_parser(self.__path_to_file)
        self.data = None
        self.__file_init()
        self.root.destroy()
        return self

    def folder_selector(self, path=None):
        if path != None:
            self.__directory = path
            return self.__directory
        self.__loger("Пользовательский выбор директории")
        sleep(3)
        self.root = tkinter.Tk()
        self.__directory = fd.askdirectory(parent=self.root, title="Куда скидывать файлы?")
        self.__loger(f'{self.__directory}')
        self.root.destroy()
        return self.__directory

    def save_file(self):
        self.__loger("Пользовательское сохранение файла")
        sleep(3)
        self.root = tkinter.Tk()
        self._new_file_name = fd.asksaveasfilename(parent=self.root)
        self.__loger(self._new_file_name)
        self.root.destroy()
        return self._new_file_name

    def __dir_parser(self, path):
        info = {}
        self.__loger(f'Анализ пути {path}')
        path=path.split("/")
        file_info = path[len(path) - 1]
        if len(path) > 1:
            path.pop(len(path) - 1)
            if "." in file_info:
                info[f"full_file_name"] = file_info
                file_info = file_info.split(".")
                info["file_type"] = file_info[-1]
                file_info.pop(-1)
                tmp = ""
                for i in file_info:
                    tmp = tmp + i
                info["file_name"] = tmp
        tmp = ""
        for i in path:
            tmp = str(tmp) + str(i) + "/"
        self.info = info

    def __csv_reader(self):
        try:
            self.__loger(f'[CSV_READER] - > Чтение файла {self.__path_to_file}')
            with open(self.__path_to_file, "r") as file:
                file_reader = csv.reader(file, delimiter=";")
                self.__loger(f'[CSV_READER] - > Парсинг строк')
                self.data = []
                for row in file_reader:
                    tmp = []
                    for i in row:
                        tmp.append(i)
                    self.data.append(tmp)
            self.__loger(f'[CSV_READER] - > Чтение завершено успешно. Прочитано {len(self.data)} строк')
            return self.data
        except Exception as e:
            self.__loger(f'[CSV_READER] - > ОШИБКА {str(e)}')
            self.__loger(f'[CSV_READER] - > Чтение завершено с ОШИБКОЙ. Прочитано {len(self.data)} строк')
            return str(e)

    def __xl_reader(self):
        letters=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z",
                 "AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ"]
        try:
            self.__loger(f'[EXCEL_READER] - > Чтение файла {self.__path_to_file}')
            xl = pd.ExcelFile(self.__path_to_file)
            self.__loger(f'[EXCEL_READER] - > Найлены листы - {xl.sheet_names}')
            wb_obj = openpyxl.load_workbook(self.__path_to_file, data_only=True)
            self.data = {}
            for i in xl.sheet_names:
                rows = []
                self.__loger(f'[EXCEL_READER] - > Чтение листа - {i}')
                sheet_obj = wb_obj[i]
                row = 1
                rowbroker = 100
                while rowbroker != 0:
                    count = 1
                    tmp = {}
                    cellbroker = 10
                    for cels in letters:
                        if sheet_obj.cell(row=row, column=count).value == None:
                            cellbroker -= 1
                        else:
                            cellbroker = 10
                        if cellbroker == 0:
                            for i in letters:
                                if tmp.get(i) == None and i in tmp.keys():
                                    tmp.pop(i)
                            break
                        tmp[cels] = sheet_obj.cell(row=row, column=count).value
                        count +=1
                    rows.append((tmp))
                    if len(tmp) == 0:
                        rowbroker -=1
                    else:
                        rowbroker = 100
                    if rowbroker == 0:
                        for i in range(0, 100):
                            rows.pop(-1)
                    row +=1
                self.data[i] = rows
                self.__loger(f'[EXCEL_READER] - > прочитано {len(rows)} строк')
            self.__loger(f'[EXCEL_READER] - > Чтение завершено успешно. Прочитано {len(self.data)} листов')
            xl.close()
            return self.data
        except Exception as e:
            self.__loger(f'[EXCEL_READER] - > ОШИБКА {e}')
            #self.__loger(f'[EXCEL_READER] - > Чтение завершено с ОШИБКОЙ. Прочитано {len(self.data)} строк')
            return str(e)
    def read_file(self, path_to_file=None, file_encoding="utf_8_sig"):
        self._file_encoding = file_encoding
        if self._file_encoding == None:
            self._file_encoding = "utf_8_sig"
        self.data = None
        if path_to_file == None:
            self.__path_to_file = self.file_selector()
        else:
            self.__path_to_file = path_to_file
            self.__dir_parser(self.__path_to_file)
            self.__file_init()
        return self

    def close(self):
        self.root.destroy()
        del self.root
    def __loger(self, *fp):
        if self.logs == True:
            if self.logtype == "print":
                print(*fp)
            if self.logtype == "file":
                logs = f'{fp}...{datetime.now()}\n'
                try:
                    with open(self._path_to_log, "r+", encoding="utf_8_sig") as f:
                        f.write(f'{fp}\n')
                except Exception as e:
                    print(e)
                    with open(self._path_to_log, "w+", encoding="utf_8_sig") as f:
                        f.write(f'{fp}\n')

#t = Files(logs=True, logtype="print")
#print(t.save_file())
#t.read_file()
#print(t.data)